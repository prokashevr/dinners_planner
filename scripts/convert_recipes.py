#!/usr/bin/env python3
"""Convert `recipies database.xlsx` -> `docs/recipes.json`.

Run from repo root:  python3 scripts/convert_recipes.py
"""
from __future__ import annotations

import datetime as dt
import json
import re
import sys
import unicodedata
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "recipies database.xlsx"
OUT = ROOT / "docs" / "recipes.json"

PROTEIN_FIXES = {
    "chiken": "chicken",
    "hering": "herring",
    "none": "vegetarian",
    "haloumi": "halloumi",
}

COMPLEXITY_MAP = {
    "ok": "simple",
    "easy": "simple",
    "simple": "simple",
    "middle": "medium",
    "medium": "medium",
    "hard": "complex",
    "complex": "complex",
    "hard and long": "complex",
}

SPLIT_PROTEIN = re.compile(r"\s+(?:and|or|/|,)\s+", re.IGNORECASE)


def slugify(text: str) -> str:
    norm = unicodedata.normalize("NFKD", text)
    ascii_part = norm.encode("ascii", "ignore").decode("ascii").strip()
    if not ascii_part:
        ascii_part = "recipe-" + str(abs(hash(text)) % 10_000)
    slug = re.sub(r"[^a-zA-Z0-9]+", "-", ascii_part).strip("-").lower()
    return slug or "recipe"


def parse_proteins(raw: str) -> list[str]:
    raw = (raw or "").strip().lower()
    if not raw:
        return ["vegetarian"]
    parts = [p.strip() for p in SPLIT_PROTEIN.split(raw) if p.strip()]
    if not parts:
        parts = [raw]
    fixed = []
    for p in parts:
        p = p.strip().rstrip(".").strip()
        p = PROTEIN_FIXES.get(p, p)
        if p == "various":
            p = "mixed"
        if p not in fixed:
            fixed.append(p)
    return fixed


def parse_complexity(raw: str) -> str:
    key = (raw or "").strip().lower()
    return COMPLEXITY_MAP.get(key, "simple")


def parse_days(value) -> int:
    """Normalize the 'for how many days' cell.

    Excel sometimes auto-converts strings like '2/3' or '1/2' into a date with
    the locale's day/month order. We decode by inspecting the (month, day) pair
    and pick the lower bound of the range.
    """
    if value is None:
        return 1
    if isinstance(value, (dt.datetime, dt.date)):
        nums = sorted({value.month, value.day})
        return max(1, min(nums))
    if isinstance(value, (int, float)):
        return max(1, int(value))
    s = str(value).strip()
    if not s:
        return 1
    m = re.match(r"\s*(\d+)\s*[-/]\s*(\d+)\s*", s)
    if m:
        return max(1, min(int(m.group(1)), int(m.group(2))))
    m = re.match(r"\s*(\d+)\s*", s)
    if m:
        return max(1, int(m.group(1)))
    return 1


def parse_author(raw: str) -> str:
    val = (raw or "").strip()
    if not val:
        return "Roma"
    return val[:1].upper() + val[1:].lower()


def main() -> int:
    if not XLSX.exists():
        print(f"missing: {XLSX}", file=sys.stderr)
        return 1
    wb = load_workbook(XLSX, data_only=True)
    ws = wb[wb.sheetnames[0]]

    recipes: list[dict] = []
    used_ids: set[str] = set()
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        print("empty sheet", file=sys.stderr)
        return 1

    for row in rows[1:]:
        name = row[0]
        if name is None or not str(name).strip():
            continue
        name = str(name).strip()
        proteins = parse_proteins(str(row[1]) if row[1] is not None else "")
        complexity = parse_complexity(str(row[2]) if row[2] is not None else "")
        days = parse_days(row[3])
        author = parse_author(str(row[4]) if row[4] is not None else "")

        base_id = slugify(name)
        rid = base_id
        i = 2
        while rid in used_ids:
            rid = f"{base_id}-{i}"
            i += 1
        used_ids.add(rid)

        recipes.append(
            {
                "id": rid,
                "name": name,
                "proteins": proteins,
                "complexity": complexity,
                "daysCovered": days,
                "author": author,
            }
        )

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(
        json.dumps(recipes, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )

    print(f"wrote {len(recipes)} recipes -> {OUT}")
    print(f"  authors:    {Counter(r['author'] for r in recipes)}")
    print(f"  complexity: {Counter(r['complexity'] for r in recipes)}")
    print(f"  days:       {Counter(r['daysCovered'] for r in recipes)}")
    proteins_flat = Counter(p for r in recipes for p in r["proteins"])
    print(f"  proteins:   {dict(proteins_flat)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
